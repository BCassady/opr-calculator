import numpy as np
import xlrd
import openpyxl
import os.path, time 

filename = "./Arkansas Champ Data.xlsx"

lastchange = os.path.getmtime(filename)

while True: 
    if lastchange < os.path.getmtime(filename):
        print("File has been changed, updating OPR")

        loc = (filename)

        data = xlrd.open_workbook(loc)

        teams = data.sheet_by_index(0)
        team_num = teams.nrows

        matches = data.sheet_by_index(1)
        match_num = matches.nrows

        index_team = {}
        team_index = {}

        i = 0
        for row in range(team_num):
            team = teams.cell_value(row, 0)
            team_index[int(team)] = i
            index_team[i] = int(team )
            i += 1


        mat = [[0]*team_num for _ in range(team_num)]
        s = [0 for _ in range(team_num)]

        for row in range(match_num):
            
            team0 = team_index[int(matches.cell_value(row, 0))]
            team1 = team_index[int(matches.cell_value(row, 1))]
            team2 = team_index[int(matches.cell_value(row, 2))]
            team3 = team_index[int(matches.cell_value(row, 3))]

            mat[team0][team0] += 1
            mat[team1][team1] += 1
            mat[team2][team2] += 1
            mat[team3][team3] += 1

            mat[team0][team1] += 1
            mat[team1][team0] += 1
            mat[team2][team3] += 1
            mat[team3][team2] += 1

            s[team0] += int(matches.cell_value(row, 4))
            s[team1] += int(matches.cell_value(row, 4))
            s[team2] += int(matches.cell_value(row, 5))
            s[team3] += int(matches.cell_value(row, 5))

        p = np.linalg.solve(mat, s)

        oprs = []

        for i in range(team_num):
            curr_team = index_team[i]
            oprs.append((curr_team, p[i]))

        oprs.sort(key = lambda x: x[1])

        oprs = oprs[::-1]


        wb = openpyxl.load_workbook(filename)

        opr_sheet = wb.worksheets[2]

        for i, opr in enumerate(oprs):
            opr_sheet.cell(row=i+1, column=1).value = opr[0]
            opr_sheet.cell(row=i+1, column=2).value = opr[1]

        wb.save(filename)
        lastchange = os.path.getmtime(filename)
    else:
        print("No file changes, waiting 3 seconds")
    time.sleep(3)